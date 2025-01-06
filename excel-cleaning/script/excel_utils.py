# excel_utils.py
from typing import Dict, List, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def load_excel_worksheet(file_path: str) -> Worksheet:
    """
    Load an Excel workbook and return its active worksheet.

    Args:
        file_path: Path to the Excel file

    Returns:
        Active worksheet from the workbook
    """
    workbook = load_workbook(file_path)
    return workbook.active


def initialize_column_headers(worksheet: Worksheet) -> Dict[int, List[str]]:
    """
    Initialize empty column headers dictionary.

    Args:
        worksheet: Excel worksheet

    Returns:
        Dictionary with column numbers as keys and empty lists as values
    """
    return {col: [] for col in range(1, worksheet.max_column + 1)}


def process_merged_cells(worksheet: Worksheet, column_headers: Dict[int, List[str]]) -> int:
    """
    Process merged cells and update column headers.

    Args:
        worksheet: Excel worksheet
        column_headers: Dictionary to store header values

    Returns:
        Maximum row number containing merged cells
    """
    max_row = 1

    for merged_range in worksheet.merged_cells.ranges:
        value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
        value = "" if value is None else str(value)

        for col in range(merged_range.min_col, merged_range.max_col + 1):
            # Fill empty cells before the merged cell
            while len(column_headers[col]) < merged_range.min_row - 1:
                column_headers[col].append("")
            column_headers[col].append(value)

        max_row = max(max_row, merged_range.max_row)

    return max_row


def process_non_merged_cells(worksheet: Worksheet, column_headers: Dict[int, List[str]], max_row: int) -> None:
    """
    Process non-merged cells and update column headers.

    Args:
        worksheet: Excel worksheet
        column_headers: Dictionary to store header values
        max_row: Maximum row number to process
    """
    for col in range(1, worksheet.max_column + 1):
        for row in range(1, max_row + 2):
            cell = worksheet.cell(row, col)

            # Skip merged cells
            if not any(cell.coordinate in merged_range
                      for merged_range in worksheet.merged_cells.ranges):
                # Fill empty cells before current cell
                while len(column_headers[col]) < row - 1:
                    column_headers[col].append("")

                value = cell.value if cell.value is not None else ""
                column_headers[col].append(str(value))


def combine_headers(column_headers: Dict[int, List[str]], max_column: int) -> List[str]:
    """
    Combine header parts for each column into final headers.

    Args:
        column_headers: Dictionary containing header values
        max_column: Maximum column number

    Returns:
        List of combined header strings
    """
    final_headers = []
    for col in range(1, max_column + 1):
        header_parts = [part for part in column_headers[col] if part]
        final_headers.append(" ".join(header_parts))
    return final_headers


def process_dataframe(file_path: str, max_row: int, final_headers: List[str]) -> pd.DataFrame:
    """
    Process Excel data into a cleaned DataFrame.

    Args:
        file_path: Path to the Excel file
        max_row: Number of rows to skip (header rows)
        final_headers: List of column headers

    Returns:
        Cleaned DataFrame with proper headers
    """
    df = pd.read_excel(file_path, header=None, skiprows=max_row)
    df.columns = final_headers
    return df.dropna(how='all').reset_index(drop=True)


def save_dataframe(df: pd.DataFrame, file_path: str) -> None:
    """
    Save DataFrame back to Excel file.

    Args:
        df: DataFrame to save
        file_path: Path where to save the file
    """
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, index=False)


def print_headers(headers: List[str], file_path: str) -> None:
    """
    Print processing results and headers.

    Args:
        headers: List of column headers
        file_path: Path to the processed file
    """
    print(f"File has been updated successfully: {file_path}")
    print("\nUpdated column headers:")
    for header in headers:
        print(f"- {header}")
