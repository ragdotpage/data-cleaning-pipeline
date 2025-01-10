import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from decorator import calculate_time

@calculate_time
def process_excel_file(input_file, output_file):
    # Load workbook to handle merged cells
    wb = load_workbook(input_file)
    ws = wb.active

    # Create a dictionary to store header values for each column
    column_headers = {}
    max_row = 1

    # Initialize column headers dictionary
    for col in range(1, ws.max_column + 1):
        column_headers[col] = []

    # Process merged cells first
    for merged_range in ws.merged_cells.ranges:
        value = ws.cell(merged_range.min_row, merged_range.min_col).value
        if value is None:
            value = ""

        # Apply the merged cell value to all affected columns
        for col in range(merged_range.min_col, merged_range.max_col + 1):
            # Ensure the list has enough elements
            while len(column_headers[col]) < merged_range.min_row - 1:
                column_headers[col].append("")
            column_headers[col].append(value)

        max_row = max(max_row, merged_range.max_row)

    # Process non-merged cells
    for col in range(1, ws.max_column + 1):
        for row in range(1, max_row + 2):  # +2 to include the row after merged cells
            cell = ws.cell(row, col)
            if not any(cell.coordinate in merged_range for merged_range in ws.merged_cells.ranges):
                # Ensure the list has enough elements
                while len(column_headers[col]) < row - 1:
                    column_headers[col].append("")
                value = cell.value if cell.value is not None else ""
                column_headers[col].append(str(value))

    # Combine headers for each column
    final_headers = []
    for col in range(1, ws.max_column + 1):
        # Filter out empty strings and combine with spaces
        header_parts = [part for part in column_headers[col] if part]
        final_headers.append(" ".join(header_parts))

    # Read the data portion of the Excel file
    df = pd.read_excel(input_file, header=None, skiprows=max_row)

    # Set the combined headers
    df.columns = final_headers

    # Remove completely empty rows
    df_cleaned = df.dropna(how='all')

    # Reset the index after removing rows
    df_cleaned = df_cleaned.reset_index(drop=True)

    # Save the cleaned DataFrame to a new Excel file
    df_cleaned.to_excel(output_file, index=False)
    print(f"File processed and saved as: {output_file}")

    # Print the headers for verification
    print("\nColumn headers:")
    for header in final_headers:
        print(f"- {header}")

if __name__ == "__main__":
    input_file = "inventory.xlsx"  # Replace with your input file name
    output_file = "inventory_cleaned_pandas.xlsx"  # Replace with your desired output file name

    try:
        process_excel_file(input_file, output_file)
    except Exception as e:
        print(f"An error occurred: {str(e)}")
